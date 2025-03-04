import sys
import os
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QLabel,
    QLineEdit,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QMessageBox,
)
from PyQt5.QtCore import Qt
import openpyxl
from PyQt5.QtGui import QIcon

EXCEL_FILE = "Book.xlsx"
BULLETINS_FOLDER = "Bulletins"


def fetch_grades(matricule):
    """Fetch grades for a given matricule from the Excel file."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        # Find the row for the given matricule in column A
        matricule_row = None
        for row in ws.iter_rows(min_col=1, max_col=1):  # Search only in column A
            cell_value = row[0].value
            # Convert both the cell value and matricule to strings for comparison
            if str(cell_value) == str(matricule):
                matricule_row = row[0].row
                break

        if not matricule_row:
            return None  # Matricule not found

        # Fetch the name and grades
        name = ws.cell(row=matricule_row, column=2).value  # Assuming name is in column B
        grades = {}
        subjects = ["Maths", "Sciences", "Histoire", "Geo", "ECM", "Francais", "Anglais"]
        for col_num, subject in enumerate(subjects, start=3):  # Grades start from column C
            grade = ws.cell(row=matricule_row, column=col_num).value
            grades[subject] = grade

        # Calculate the average
        valid_grades = [grade for grade in grades.values() if isinstance(grade, (int, float))]
        if valid_grades:
            grades["Moyenne"] = sum(valid_grades) / len(valid_grades)
        else:
            grades["Moyenne"] = "N/A"

        return {"name": name, "grades": grades}
    except Exception as e:
        print(f"Error fetching grades: {e}")
        return None


def create_new_excel_file():
    """Create a new Excel file in the Bulletins folder."""
    try:
        # Create the Bulletins folder if it doesn't exist
        if not os.path.exists(BULLETINS_FOLDER):
            os.makedirs(BULLETINS_FOLDER)

        # Create a new Excel file
        file_name = f"Bulletin_{len(os.listdir(BULLETINS_FOLDER)) + 1}.xlsx"
        file_path = os.path.join(BULLETINS_FOLDER, file_name)

        # Create a new workbook and add a default sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grades"

        # Add headers
        headers = ["Matricule", "Name", "Maths", "Sciences", "Histoire", "Geo", "ECM", "Francais", "Anglais"]
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        # Save the workbook
        wb.save(file_path)
        return file_path
    except Exception as e:
        print(f"Error creating new Excel file: {e}")
        return None


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("View Student Grades")
        self.setGeometry(100, 100, 1920, 1080)
        self.setWindowIcon(QIcon('icon.png'))

        # Main widget and layout
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        # Matricule input field
        self.matricule_field = QLineEdit(self)
        self.matricule_field.setPlaceholderText("Enter Matricule")
        self.layout.addWidget(self.matricule_field)

        # View Grades button
        self.view_button = QPushButton("View Grades", self)
        self.view_button.clicked.connect(self.view_grades)
        self.layout.addWidget(self.view_button)

        # Create New Excel File button
        self.create_button = QPushButton("Create New Excel File", self)
        self.create_button.clicked.connect(self.create_excel_file)
        self.layout.addWidget(self.create_button)

        # Result label for messages
        self.result_label = QLabel(self)
        self.result_label.setAlignment(Qt.AlignCenter)
        self.layout.addWidget(self.result_label)

        # Table to display grades
        self.grades_table = QTableWidget(self)
        self.grades_table.setColumnCount(2)
        self.grades_table.setHorizontalHeaderLabels(["Subject", "Grade"])
        self.grades_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.layout.addWidget(self.grades_table)

    def view_grades(self):
        matricule = self.matricule_field.text().strip()
        if not matricule:
            self.result_label.setText("Please enter a Matricule!")
            self.result_label.setStyleSheet("color: red;")
            return

        # Fetch grades from Excel
        data = fetch_grades(matricule)
        if not data:
            self.result_label.setText(f"No data found for Matricule: {matricule}")
            self.result_label.setStyleSheet("color: red;")
            self.grades_table.setRowCount(0)  # Clear the table
        else:
            self.result_label.setText(f"Grades for Matricule: {matricule}, Name: {data['name']}")
            self.result_label.setStyleSheet("color: green;")

            # Populate the table with grades
            self.grades_table.setRowCount(len(data["grades"]))
            for row, (subject, grade) in enumerate(data["grades"].items()):
                self.grades_table.setItem(row, 0, QTableWidgetItem(subject))
                self.grades_table.setItem(row, 1, QTableWidgetItem(str(grade)))

    def create_excel_file(self):
        """Create a new Excel file and show a success message."""
        file_path = create_new_excel_file()
        if file_path:
            QMessageBox.information(self, "Success", f"New Excel file created at:\n{file_path}")
        else:
            QMessageBox.critical(self, "Error", "Failed to create a new Excel file.")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())