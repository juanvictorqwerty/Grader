import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QWidget, QFormLayout, QLineEdit, QMessageBox
from PyQt5.QtGui import QIcon
from openpyxl import load_workbook

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Grader")
        self.setGeometry(100, 100, 1920, 1080)
        self.setWindowIcon(QIcon('icon.png'))
        
        # Create a central widget and set the layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        form_layout = QFormLayout()
        central_widget.setLayout(form_layout)
        
        # Create input fields
        self.MatriculeInput = QLineEdit(self)
        form_layout.addRow("Enter your matric:", self.MatriculeInput)
        
        self.NameInput = QLineEdit(self)
        form_layout.addRow("Enter your name:", self.NameInput)

        self.MathsInput = QLineEdit(self)
        form_layout.addRow("Enter your Maths score:", self.MathsInput)

        self.ScienceInput = QLineEdit(self)
        form_layout.addRow("Enter your Science score:", self.ScienceInput)

        self.EnglishInput = QLineEdit(self)
        form_layout.addRow("Enter your English score:", self.EnglishInput)

        self.FrancaisInput = QLineEdit(self)
        form_layout.addRow("Enter your Francais score:", self.FrancaisInput)  

        self.HistoireInput = QLineEdit(self)
        form_layout.addRow("Enter your Histoire score:", self.HistoireInput)

        self.GeoInput = QLineEdit(self)
        form_layout.addRow("Enter your Geo score:", self.GeoInput)

        # Create and connect the submit button
        submit_button = QPushButton("Submit")
        submit_button.clicked.connect(self.on_submit)
        form_layout.addRow(submit_button)
        
        # Create and connect the back button
        self.GoBack = QPushButton('Back', self)
        self.GoBack.clicked.connect(self.go_home)
        form_layout.addRow(self.GoBack)
        
    def on_submit(self):
        # Collect data from input fields
        matric = self.MatriculeInput.text()
        name = self.NameInput.text()
        maths = self.MathsInput.text()
        science = self.ScienceInput.text()
        english = self.EnglishInput.text()
        francais = self.FrancaisInput.text()
        histoire = self.HistoireInput.text()
        geo = self.GeoInput.text()
        
        # Validate inputs (basic validation)
        if not all([matric, name, maths, science, english, francais, histoire, geo]):
            QMessageBox.warning(self, "Input Error", "Please fill in all fields.")
            return
        
        # Convert scores to floats for calculation
        try:
            maths = float(maths)
            science = float(science)
            english = float(english)
            francais = float(francais)
            histoire = float(histoire)
            geo = float(geo)
        except ValueError:
            QMessageBox.warning(self, "Input Error", "All scores must be numeric.")
            return
        
        # Calculate the average (Moyenne)
        moyenne = (maths + science + english + francais + histoire + geo) / 6
        
        # Append data to Excel file
        try:
            # Load the workbook and select the sheet
            workbook = load_workbook("Book.xlsx")
            sheet = workbook["Feuille 1"]
            
            # Find the next empty row
            next_row = sheet.max_row + 1
            
            # Append data to specific columns
            sheet.cell(row=next_row, column=1, value=matric)  # Column A: Matric
            sheet.cell(row=next_row, column=2, value=name)    # Column B: Name
            sheet.cell(row=next_row, column=3, value=maths)   # Column C: Maths
            sheet.cell(row=next_row, column=4, value=science) # Column D: Science
            sheet.cell(row=next_row, column=5, value=english) # Column E: English
            sheet.cell(row=next_row, column=6, value=francais) # Column F: Francais
            sheet.cell(row=next_row, column=7, value=histoire) # Column G: Histoire
            sheet.cell(row=next_row, column=8, value=geo)      # Column H: Geo
            sheet.cell(row=next_row, column=9, value=moyenne)  # Column I: Moyenne
            
            # Save the workbook
            workbook.save("Book.xlsx")
            
            QMessageBox.information(self, "Submission Successful", "Data has been saved to Book.xlsx.")
        except FileNotFoundError:
            QMessageBox.critical(self, "Error", "The file 'Book.xlsx' was not found.")
        except PermissionError:
            QMessageBox.critical(self, "Error", "Permission denied. Please ensure the file is not open in another program.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred while saving the data: {str(e)}")
        
    def go_home(self):
        # Reset the form or navigate back (you can customize this)
        self.MatriculeInput.clear()
        self.NameInput.clear()
        self.MathsInput.clear()
        self.ScienceInput.clear()
        self.EnglishInput.clear()
        self.FrancaisInput.clear()
        self.HistoireInput.clear()
        self.GeoInput.clear()
        QMessageBox.information(self, "Info", "Form reset.")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())